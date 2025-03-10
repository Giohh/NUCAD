from playwright.sync_api import sync_playwright, TimeoutError
import time
import logging
from openpyxl import Workbook

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def extrair_siafi():   
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.connect_over_cdp('http://localhost:9230', slow_mo=600)

            # Recuperar o contexto padrão e a primeira página
            default_context = browser.contexts[0]
            pages = default_context.pages
            page = pages[0]
            
            logger.info("Conectado ao navegador")
            
            # Verificar se o iframe específico existe
            iframe_selector = "iframe[id='iframe'] >> nth=0"
            iframe_element = page.query_selector(iframe_selector)
            
            if iframe_element:
                logger.info("✅ iFrame específico encontrado!")
                
                # Obter propriedades do iframe para confirmar
                iframe_id = iframe_element.get_attribute('id') or 'Sem ID'
                iframe_name = iframe_element.get_attribute('name') or 'Sem name'
                iframe_src = iframe_element.get_attribute('src') or 'Sem src'
                
                logger.info(f"Propriedades do iframe: id='{iframe_id}', name='{iframe_name}', src='{iframe_src}'")
                
                # Obter o frame correspondente a este iframe
                iframe_frame = None
                for frame in page.frames:
                    if frame.name == iframe_id or frame.name == iframe_name:
                        iframe_frame = frame
                        logger.info(f"Frame encontrado pelo nome: {frame.name}")
                        break
                
                # Se não encontrou pelo nome, tentar pelo URL
                if not iframe_frame and iframe_src:
                    for frame in page.frames:
                        if iframe_src in frame.url:
                            iframe_frame = frame
                            logger.info(f"Frame encontrado pela URL: {frame.url}")
                            break
                
                # Se ainda não encontrou, usar uma abordagem mais genérica
                if not iframe_frame:
                    logger.warning("Frame específico não encontrado. Tentando abordar de outra forma...")
                    
                    # Tentar obter o frame diretamente pelo Playwright
                    iframe_frame = page.frame(iframe_id) or page.frame(iframe_name)
                    
                    if iframe_frame:
                        logger.info("Frame obtido via page.frame()")
                    else:
                        # Último recurso: tentar obter o conteúdo do iframe
                        logger.info("Tentando obter conteúdo do iframe...")
                        
                        # Esperar garantir que o iframe tenha carregado
                        page.wait_for_selector(iframe_selector, state="attached", timeout=10000)
                        
                        # Obter o frame usando o índice
                        all_frames = page.frames
                        if len(all_frames) > 1:
                            # Assumir que o frame que queremos não é o principal (índice 0)
                            iframe_frame = all_frames[1]  # Usar o segundo frame (índice 1)
                            logger.info(f"Usando segundo frame disponível: {iframe_frame.name}")
                        else:
                            logger.error("Não há frames suficientes disponíveis")
                            return None
            else:
                logger.error("❌ iFrame específico não encontrado")
                return None
            
            # Verificar se conseguimos um frame para trabalhar
            if not iframe_frame:
                logger.error("Não foi possível obter acesso ao conteúdo do iframe")
                return None
            
            logger.info(f"Usando frame: name='{iframe_frame.name}', url='{iframe_frame.url}'")
            
            # Agora vamos trabalhar no contexto do iframe
            try:
                # Tentar localizar e clicar no botão "Total da Lista"
                localizadores_botao = [
                    "span:has-text('Total da Lista')",
                    "span.ui-button-text.ui-clickable:has-text('Total')",
                    "button span:has-text('Total')",
                    "span[class='ui-button-text ui-clickable'] >> nth=0"
                ]
                
                botao_clicado = False
                for localizador in localizadores_botao:
                    try:
                        logger.info(f"Tentando localizador de botão: {localizador}")
                        botao = iframe_frame.query_selector(localizador)
                        
                        if botao:
                            texto_botao = botao.inner_text().strip()
                            logger.info(f"Botão encontrado: '{texto_botao}'")
                            
                            # Verificar se o botão está visível
                            visivel = botao.is_visible()
                            logger.info(f"Botão está visível? {visivel}")
                            
                            if visivel:
                                # Rolar para o botão e clicar
                                botao.scroll_into_view_if_needed()
                                botao.click(force=True)
                                logger.info("✅ Clique realizado com sucesso!")
                                botao_clicado = True
                                time.sleep(3)
                                break
                            else:
                                logger.warning("Botão encontrado mas não está visível")
                        else:
                            logger.warning(f"Botão não encontrado com localizador: {localizador}")
                    
                    except Exception as e:
                        logger.error(f"Erro ao tentar clicar com localizador {localizador}: {str(e)}")
                
                # Se não conseguiu clicar com os localizadores, tentar JavaScript
                if not botao_clicado:
                    logger.info("Tentando clicar via JavaScript...")
                    
                    resultado = iframe_frame.evaluate("""() => {
                        // Buscar por elemento com texto "Total"
                        const elementos = Array.from(document.querySelectorAll('*'));
                        const botoes = elementos.filter(el => {
                            const texto = el.textContent && el.textContent.trim();
                            return texto && texto.includes('Total');
                        });
                        
                        if (botoes.length > 0) {
                            console.log("Botão encontrado via JS:", botoes[0].textContent);
                            botoes[0].click();
                            return {sucesso: true, texto: botoes[0].textContent};
                        }
                        
                        return {sucesso: false, elementos_encontrados: elementos.length};
                    }""")
                    
                    if resultado.get('sucesso'):
                        logger.info(f"✅ Clique realizado via JavaScript: {resultado.get('texto')}")
                        botao_clicado = True
                        time.sleep(3)
                    else:
                        logger.error(f"Falha ao clicar via JavaScript. Elementos encontrados: {resultado.get('elementos_encontrados', 0)}")
                
                if not botao_clicado:
                    logger.error("Não foi possível clicar no botão por nenhum método")
                    return None
                
                # Agora vamos extrair os valores que aparecem após o clique
                logger.info("Extraindo valores...")
                
                # Localizadores para os valores - Adicionados os novos localizadores
                localizadores_valores = {
                    "Valor Incluído": "section[class='totalizador ng-star-inserted'] div[class='right ng-star-inserted'] >> nth=0",
                    "Valor Reforçado": "section[class='totalizador ng-star-inserted'] div[class='right ng-star-inserted'] >> nth=1",
                    "Valor Anulado": "section[class='totalizador ng-star-inserted'] div[class='right ng-star-inserted'] >> nth=2",
                    "Valor Atual": "section[class='totalizador ng-star-inserted'] div[class='right'] >> nth=0",
                    "Valor a Liquidar": "div[class='right'] >> nth=2",
                    "Valor em Liquidação": "div[class='right'] >> nth=3",
                    "Valor Liquidado a Pagar": "div[class='right'] >> nth=4",
                    "Valor Pago": "div[class='right'] >> nth=5"
                }
                
                # Extrair valores usando os localizadores
                valores = {}
                for nome, localizador in localizadores_valores.items():
                    try:
                        elemento = iframe_frame.query_selector(localizador)
                        if elemento:
                            valor = elemento.inner_text().strip()
                            valores[nome] = valor
                            logger.info(f"✅ {nome}: {valor}")
                        else:
                            valores[nome] = "N/A"
                            logger.warning(f"⚠️ Não foi possível encontrar {nome}")
                    except Exception as e:
                        valores[nome] = "Erro"
                        logger.error(f"❌ Erro ao extrair {nome}: {str(e)}")
                
                # Se não encontrou os valores, tentar localizadores alternativos
                if all(v == "N/A" or v == "Erro" for v in valores.values()):
                    logger.info("Tentando localizadores alternativos...")
                    
                    localizadores_alternativos = {
                        "Valor Incluído": "div.right.ng-star-inserted >> nth=0",
                        "Valor Reforçado": "div.right.ng-star-inserted >> nth=1",
                        "Valor Anulado": "div.right.ng-star-inserted >> nth=2",
                        "Valor Atual": "div.right >> nth=0",
                        "Valor a Liquidar": "div.right >> nth=2",
                        "Valor em Liquidação": "div.right >> nth=3",
                        "Valor Liquidado a Pagar": "div.right >> nth=4",
                        "Valor Pago": "div.right >> nth=5"
                    }
                    
                    for nome, localizador in localizadores_alternativos.items():
                        try:
                            elemento = iframe_frame.query_selector(localizador)
                            if elemento:
                                valor = elemento.inner_text().strip()
                                valores[nome] = valor
                                logger.info(f"✅ {nome} (alternativo): {valor}")
                        except Exception as e:
                            logger.error(f"❌ Erro ao extrair {nome} com localizador alternativo: {str(e)}")
                
                # Se ainda não encontrou, tentar via JavaScript
                if all(v == "N/A" or v == "Erro" for v in valores.values()):
                    logger.info("Tentando extrair valores via JavaScript...")
                    
                    js_valores = iframe_frame.evaluate("""() => {
                        const resultado = {};
                        
                        // Obter todos os elementos com classe 'right'
                        const rights = document.querySelectorAll('.right');
                        const rightTexts = Array.from(rights).map(el => el.textContent.trim());
                        
                        // Se encontrou elementos suficientes, usar os primeiros
                        if (rightTexts.length >= 8) {  // Atualizado para incluir os novos valores
                            resultado['Valor Incluído'] = rightTexts[0];
                            resultado['Valor Reforçado'] = rightTexts[1];
                            resultado['Valor Anulado'] = rightTexts[2];
                            resultado['Valor Atual'] = rightTexts[3];
                            resultado['Valor a Liquidar'] = rightTexts[4];
                            resultado['Valor em Liquidação'] = rightTexts[5];
                            resultado['Valor Liquidado a Pagar'] = rightTexts[6];
                            resultado['Valor Pago'] = rightTexts[7];
                            return resultado;
                        }
                        
                        // Tentar encontrar por texto adjacente
                        const divs = document.querySelectorAll('div');
                        for (let i = 0; i < divs.length; i++) {
                            const texto = divs[i].textContent.trim();
                            
                            if (texto.includes('Valor incluído') || texto.includes('(+) Valor reforçado')) {
                                // Procurar o próximo elemento que pode conter o valor
                                if (i+1 < divs.length) {
                                    const proximo = divs[i+1];
                                    if (proximo.textContent.match(/[0-9.,]+/)) {
                                        if (texto.includes('Valor incluído')) {
                                            resultado['Valor Incluído'] = proximo.textContent.trim();
                                        } else {
                                            resultado['Valor Reforçado'] = proximo.textContent.trim();
                                        }
                                    }
                                }
                            } else if (texto.includes('Valor anulado') || texto.includes('(-) Valor anulado')) {
                                if (i+1 < divs.length) {
                                    const proximo = divs[i+1];
                                    if (proximo.textContent.match(/[0-9.,]+/)) {
                                        resultado['Valor Anulado'] = proximo.textContent.trim();
                                    }
                                }
                            } else if (texto.includes('Valor atual') || texto.includes('(=) Valor atual')) {
                                if (i+1 < divs.length) {
                                    const proximo = divs[i+1];
                                    if (proximo.textContent.match(/[0-9.,]+/)) {
                                        resultado['Valor Atual'] = proximo.textContent.trim();
                                    }
                                }
                            } else if (texto.includes('Valor a liquidar')) {
                                if (i+1 < divs.length) {
                                    const proximo = divs[i+1];
                                    if (proximo.textContent.match(/[0-9.,]+/)) {
                                        resultado['Valor a Liquidar'] = proximo.textContent.trim();
                                    }
                                }
                            } else if (texto.includes('Valor em liquidação')) {
                                if (i+1 < divs.length) {
                                    const proximo = divs[i+1];
                                    if (proximo.textContent.match(/[0-9.,]+/)) {
                                        resultado['Valor em Liquidação'] = proximo.textContent.trim();
                                    }
                                }
                            } else if (texto.includes('Valor liquidado a pagar')) {
                                if (i+1 < divs.length) {
                                    const proximo = divs[i+1];
                                    if (proximo.textContent.match(/[0-9.,]+/)) {
                                        resultado['Valor Liquidado a Pagar'] = proximo.textContent.trim();
                                    }
                                }
                            } else if (texto.includes('Valor pago')) {
                                if (i+1 < divs.length) {
                                    const proximo = divs[i+1];
                                    if (proximo.textContent.match(/[0-9.,]+/)) {
                                        resultado['Valor Pago'] = proximo.textContent.trim();
                                    }
                                }
                            }
                        }
                        
                        return resultado;
                    }""")
                    
                    # Atualizar valores encontrados via JavaScript
                    for nome, valor in js_valores.items():
                        valores[nome] = valor
                        logger.info(f"✅ {nome} (via JS): {valor}")
                
                # Criar planilha para salvar os dados
                wb = Workbook()
                ws = wb.active
                
                # Cabeçalhos
                for coluna, nome in enumerate(valores.keys(), start=1):
                    ws.cell(row=1, column=coluna, value=nome)
                
                # Valores
                for coluna, valor in enumerate(valores.values(), start=1):
                    ws.cell(row=2, column=coluna, value=valor)
                
                # Salvar planilha
                wb.save("valores_siafi.xlsx")
                logger.info("Dados salvos na planilha 'valores_siafi.xlsx'")
                
                return valores
            
            except Exception as e:
                logger.error(f"Erro ao interagir com o iframe: {str(e)}")
                return None
    
    except Exception as e:
        logger.error(f"Erro geral na execução: {str(e)}")
        return None

# Executar o código
if __name__ == "__main__":
    extrair_siafi()